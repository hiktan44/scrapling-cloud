"""Toplu site zekası: Excel/CSV/TXT yüklemesinden URL çıkarma, sosyal medya
bağlantısı tespiti, kısa firma analizi ve Excel export üretimi."""

from __future__ import annotations

import csv
import io
import re

from .analyzer import extract_with_prompt

MAX_URLS = 1000
MAX_ROWS = 50000

URL_RE = re.compile(r"https?://[^\s,\"'<>\)\]]+", re.IGNORECASE)

# Google Maps bağlantıları hiçbir zaman firma sitesi değildir
MAPS_RE = re.compile(r"^https?://(www\.)?google\.[a-z.]+/maps", re.IGNORECASE)

# Bu başlıkları taşıyan sütunlar site listesi sayılır. Böyle bir sütun varsa
# yalnızca ondan URL alınır (Google Maps / sosyal medya sütunları atlanır).
WEBSITE_HEADERS = {
    "website", "web site", "web", "site", "url", "domain",
    "web sitesi", "websitesi", "web adresi", "site adresi", "internet sitesi",
}


def _is_website_header(value: str) -> bool:
    v = (value or "").strip().lower().replace("i̇", "i")
    return v in WEBSITE_HEADERS or v.startswith(("website", "web site", "web sitesi", "websitesi"))

SOCIAL_DOMAINS: list[tuple[str, tuple[str, ...]]] = [
    ("instagram", ("instagram.com",)),
    ("facebook", ("facebook.com", "fb.com", "fb.me")),
    ("x", ("twitter.com", "x.com")),
    ("linkedin", ("linkedin.com",)),
    ("youtube", ("youtube.com", "youtu.be")),
    ("tiktok", ("tiktok.com",)),
    ("whatsapp", ("wa.me", "whatsapp.com")),
    ("telegram", ("t.me", "telegram.me")),
    ("pinterest", ("pinterest.",)),
    ("github", ("github.com",)),
]

# Profil sayfası olmayan paylaşım/yardımcı bağlantılarını ele
SOCIAL_BAD_PATHS = ("sharer", "/share", "/intent/", "/plugins/", "/dialog/", "hashtag", "/embed/", "/watch")

INTEL_PROMPT = (
    "Sen bir B2B pazar araştırması analistisin. Aşağıdaki web sitesi içeriğini incele ve "
    "şirketin hangi ürünleri sattığını veya hangi hizmetleri sunduğunu TÜRKÇE, en fazla 2-3 "
    "cümleyle özetle. Genel ifadelerden kaçın; somut ürün, hizmet veya sektör adları kullan. "
    'Yalnızca şu formatta JSON döndür: {"analiz": "..."}'
)


def _normalize_url(value: str) -> str | None:
    u = (value or "").strip().strip("\"\'").strip()
    if not u or "@" in u or " " in u:
        return None
    low = u.lower()
    if low.startswith(("http://", "https://")):
        pass
    elif "." in u and not low.startswith(("mailto:", "tel:", "javascript:")):
        u = "https://" + u
    else:
        return None
    return u.rstrip("/.,;")


# Bu hostlarda path firmayı belirtir (instagram.com/a ile instagram.com/b farklı firmalar)
PATH_SENSITIVE_HOSTS = {
    "instagram.com", "facebook.com", "twitter.com", "x.com", "linkedin.com",
    "youtube.com", "youtu.be", "tiktok.com", "wa.me", "t.me", "telegram.me",
    "github.com", "linktr.ee", "threads.net",
}


def _add_url(found: list[str], seen: set[str], raw: str) -> None:
    u = _normalize_url(raw)
    if not u:
        return
    if MAPS_RE.match(u):
        return
    host = re.sub(r"^https?://(www\.)?", "", u.lower()).split("/")[0]
    if not host or "." not in host:
        return
    if host in PATH_SENSITIVE_HOSTS or any(host.endswith("." + h) for h in PATH_SENSITIVE_HOSTS):
        seg = re.sub(r"^https?://[^/]+", "", u.lower()).strip("/").split("/")[0]
        key = f"{host}/{seg}" if seg else host
    else:
        key = host  # aynı domaini bir kez işle
    if key in seen:
        return
    seen.add(key)
    found.append(u)


def parse_urls(filename: str, content: bytes) -> list[str]:
    """Yüklenen dosyadan benzersiz URL listesi çıkar (xlsx, csv veya txt).

    Dosyada "Website" benzeri bir başlık sütunu varsa yalnızca o sütun(lar)
    kullanılır (firma rehberi modu); yoksa tüm hücreler/satırlar taranır.
    """
    name = (filename or "").lower()
    column_urls: list[str] = []   # website sütunundan gelenler (öncelikli)
    column_seen: set[str] = set()
    loose_urls: list[str] = []    # serbest taramadan gelenler (yedek)
    loose_seen: set[str] = set()

    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise ValueError("Excel desteği için openpyxl gerekli") from exc
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError("Excel dosyası okunamadı (bozuk veya şifreli olabilir)") from exc
        try:
            for sheet in wb.worksheets:
                web_cols: list[int] = []
                for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    if r_idx >= MAX_ROWS or len(column_urls) >= MAX_URLS:
                        break
                    if r_idx <= 4 and not web_cols:
                        web_cols = [i for i, cell in enumerate(row) if cell is not None and _is_website_header(str(cell))]
                    if web_cols:
                        for i in web_cols:
                            if i < len(row) and row[i] is not None:
                                text = str(row[i])
                                for m in URL_RE.findall(text) or [text]:
                                    _add_url(column_urls, column_seen, m)
                    else:
                        for cell in row:
                            if cell is None:
                                continue
                            text = str(cell)
                            for m in URL_RE.findall(text) or [text]:
                                _add_url(loose_urls, loose_seen, m)
        finally:
            wb.close()
    else:
        try:
            text = content.decode("utf-8-sig", errors="replace")
        except Exception as exc:
            raise ValueError("Dosya metni okunamadı") from exc
        if name.endswith(".csv"):
            try:
                rows = csv.reader(io.StringIO(text))
                web_cols: list[int] = []
                for r_idx, row in enumerate(rows):
                    if r_idx >= MAX_ROWS or len(column_urls) >= MAX_URLS:
                        break
                    if r_idx <= 4 and not web_cols:
                        web_cols = [i for i, cell in enumerate(row) if _is_website_header(cell)]
                    if web_cols:
                        for i in web_cols:
                            if i < len(row):
                                _add_url(column_urls, column_seen, row[i])
                    else:
                        for cell in row:
                            _add_url(loose_urls, loose_seen, cell)
            except csv.Error:
                for line in text.splitlines():
                    _add_url(loose_urls, loose_seen, line)
        else:
            for line in text.splitlines()[:MAX_ROWS]:
                for m in URL_RE.findall(line) or [line]:
                    _add_url(loose_urls, loose_seen, m)
                if len(loose_urls) >= MAX_URLS:
                    break

    result = column_urls if column_urls else loose_urls
    return result[:MAX_URLS]


def _clean_social_url(href: str) -> str | None:
    low = href.lower().strip()
    if not low.startswith(("http://", "https://")):
        return None
    cleaned = href.split("#")[0].split("?")[0].rstrip("/")
    path = re.sub(r"^https?://[^/]+", "", cleaned.lower())
    if any(bad in path for bad in SOCIAL_BAD_PATHS):
        return None
    return cleaned


def extract_socials(links: list | None, markdown: str | None = None, site_host: str | None = None) -> dict[str, str]:
    """Sayfadaki bağlantılardan sosyal medya profil URL'lerini çıkar.

    `links` scraper çıktısı olarak düz URL stringleri (veya {"href": ...}
    sözlükleri) içerebilir; ikisi de desteklenir. `site_host` verilirse
    sitenin kendi domain'ine giden bağlantılar sosyal sayılmaz.
    """
    own = (site_host or "").lower().removeprefix("www.")
    candidates: list[str] = []
    for link in links or []:
        href = link if isinstance(link, str) else ((link or {}).get("href") or "")
        if href:
            candidates.append(href)
    if markdown:
        candidates.extend(URL_RE.findall(markdown)[:500])

    found: dict[str, str] = {}
    roots: dict[str, str] = {}
    for href in candidates:
        low = href.lower()
        host = re.sub(r"^https?://(www\.)?", "", low).split("/")[0]
        if own and (host == own or host.endswith("." + own)):
            continue
        for key, domains in SOCIAL_DOMAINS:
            if key in found:
                continue
            if any(
                host.startswith(d) if d.endswith(".") else (host == d or host.endswith("." + d))
                for d in domains
            ):
                cleaned = _clean_social_url(href)
                if not cleaned:
                    break
                path = re.sub(r"^https?://[^/]+", "", cleaned).strip("/")
                if path:
                    # Gerçek profil bağlantısı - her zaman kök bağlantıya tercih edilir
                    found[key] = cleaned
                    roots.pop(key, None)
                elif key not in roots:
                    # Platform kökü (örn. github.com) - daha iyi aday gelmezse kullanılır
                    roots[key] = cleaned
                break
    for key, url in roots.items():
        found.setdefault(key, url)
    return found


def social_for_url(url: str) -> dict[str, str]:
    """URL'nin kendisi bir sosyal medya profiliyse platform eşleşmesini döndür."""
    host = re.sub(r"^https?://(www\.)?", "", (url or "").lower()).split("/")[0]
    out: dict[str, str] = {}
    for key, domains in SOCIAL_DOMAINS:
        if any(
            host.startswith(d) if d.endswith(".") else (host == d or host.endswith("." + d))
            for d in domains
        ):
            cleaned = _clean_social_url(url)
            if cleaned:
                out[key] = cleaned
    return out


async def analyze_company(scrape_result: dict) -> str:
    """Kısa Türkçe firma analizi üret; LLM yoksa metadata'ya düş."""
    metadata = scrape_result.get("metadata") or {}
    fallback = (metadata.get("description") or metadata.get("title") or "").strip()

    trimmed = dict(scrape_result)
    md = trimmed.get("markdown") or ""
    trimmed["markdown"] = md[:12000]

    prompt_result = await extract_with_prompt(trimmed, INTEL_PROMPT)
    if isinstance(prompt_result, dict):
        analiz = prompt_result.get("analiz")
        if isinstance(analiz, str) and analiz.strip():
            return analiz.strip()[:800]
    return fallback[:400] if fallback else "Analiz üretilemedi."


def build_export_xlsx(items: list[dict]) -> bytes:
    """Sonuçları indirilebilir Excel dosyasına dönüştür."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Firma Analizi"

    headers = ["URL", "Durum", "Instagram", "Facebook", "X", "LinkedIn", "YouTube",
               "TikTok", "WhatsApp", "Telegram", "Pinterest", "GitHub", "Firma Analizi", "Hata"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(bold=True, color="FFFFFF")
    for col, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center")

    social_keys = ["instagram", "facebook", "x", "linkedin", "youtube", "tiktok",
                   "whatsapp", "telegram", "pinterest", "github"]
    for item in items:
        socials = item.get("socials") or {}
        row = [item.get("url") or "", "Başarılı" if item.get("status") == "ok" else "Hata"]
        row.extend(socials.get(k, "") for k in social_keys)
        row.append(item.get("analysis") or "")
        row.append(item.get("error") or "")
        ws.append(row)

    widths = [38, 10, 30, 30, 26, 34, 30, 28, 24, 26, 28, 28, 70, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=13).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
